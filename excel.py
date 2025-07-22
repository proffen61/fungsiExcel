import streamlit as st
from openpyxl import load_workbook
from datetime import datetime
from collections import Counter
import tempfile
import os

# ===== Utility Functions =====

def format_tgl(tgl_value):
    if isinstance(tgl_value, datetime):
        return tgl_value.strftime('%d-%m-%Y')
    elif isinstance(tgl_value, str):
        try:
            parsed = datetime.strptime(tgl_value, '%Y-%m-%d')
            return parsed.strftime('%d-%m-%Y')
        except:
            return tgl_value
    else:
        return ''

def hitung_total(masuk, keluar, harga):
    a = masuk if masuk else 0
    b = keluar if keluar else 0
    c = harga if harga else 0
    return (a * c) + (b * c)

# ===== Streamlit App =====

st.title("📘 Mutasi Excel Processor")

uploaded_file = st.file_uploader("Unggah file Excel", type=["xlsx"])

if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name

    output_path = tmp_path.replace(".xlsx", "_output.xlsx")

    mutasi_sheet_name = 'mutasi'
    source_sheets = {
        'TB': '(TB)',
        'AG': '(AG)',
        'NT': 'Notebook (NT)',
        'BY': '(BY)',
        'PT': '(PT)',
        'BK': 'tab POKA (BK)',
        'UM': 'umbul2 (UM)',
        'SP': 'Simpel (SP)',
        'TMS': 'tamasya (TMS)'
    }

    if st.button("🚀 Jalankan Proses"):
        wb = load_workbook(tmp_path)
        mutasi_ws = wb[mutasi_sheet_name]
        sheet_mutasi_counter = {sheet: Counter() for sheet in source_sheets.values()}

        # Mutasi parsing
        for row in mutasi_ws.iter_rows(min_row=3, values_only=True):
            tgl = format_tgl(row[2])
            kode_brg = row[3]
            masuk = row[4]
            keluar = row[5]
            harga = row[6]
            total = row[7]

            target_sheet = None
            if kode_brg:
                for prefix, sheet_name in source_sheets.items():
                    if prefix in kode_brg:
                        target_sheet = sheet_name
                        break
            else:
                continue
            if target_sheet is None:
                continue

            key = (tgl, kode_brg, masuk, keluar, harga, total)
            sheet_mutasi_counter[target_sheet][key] += 1

        # Inject into target sheets
        for target_sheet, counter in sheet_mutasi_counter.items():
            ws = wb[target_sheet]

            max_data_row = 2
            for idx, row in enumerate(ws.iter_rows(min_row=3), start=3):
                kode = row[2].value
                masuk = row[3].value
                keluar = row[4].value
                harga = row[5].value
                total = row[6].value

                if kode or masuk or keluar or harga or total:
                    max_data_row = idx

            last_row = ws.max_row
            if last_row > max_data_row:
                ws.delete_rows(max_data_row + 1, last_row - max_data_row)

            sheet_counter = Counter()
            for row in ws.iter_rows(min_row=3):
                tgl = format_tgl(row[1].value)
                kode = row[2].value
                masuk = row[3].value
                keluar = row[4].value
                harga = row[5].value
                total = row[6].value
                key = (tgl, kode, masuk, keluar, harga, total)
                if kode:
                    sheet_counter[key] += 1

            for key, mutasi_count in counter.items():
                sheet_count = sheet_counter.get(key, 0)
                if sheet_count < mutasi_count:
                    need_to_add = mutasi_count - sheet_count
                    for _ in range(need_to_add):
                        new_row_idx = ws.max_row + 1
                        ws[f'B{new_row_idx}'] = key[0]
                        ws[f'C{new_row_idx}'] = key[1]
                        ws[f'D{new_row_idx}'] = key[2]
                        ws[f'E{new_row_idx}'] = key[3]
                        ws[f'F{new_row_idx}'] = key[4]

                        if key[2] or key[3]:
                            ws[f'G{new_row_idx}'] = hitung_total(key[2], key[3], key[4])
                        else:
                            ws[f'G{new_row_idx}'] = f'=IF(OR(D{new_row_idx}=0,E{new_row_idx}=0),"",(D{new_row_idx}*F{new_row_idx})+(E{new_row_idx}*F{new_row_idx}))'

        # Save output
        wb.save(output_path)
        st.success("✅ File berhasil diproses!")

        with open(output_path, "rb") as f:
            st.download_button("📥 Unduh Output", f, file_name="output.xlsx")

        os.remove(output_path)
        os.remove(tmp_path)
